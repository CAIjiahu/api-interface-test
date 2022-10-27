#!C:\pythonCode
# -*- coding: utf-8 -*-
# @Time : 2022/5/5 16:35
# @Author : qiaoqiao
# @File : handle_excel.py
# @Software: PyCharm
# @python version: 3.7


import openpyxl


class HandleExcel:

    def __init__(self, filename, sheetname):
        '''

        :param filename: Excel的文件名（路径）
        :param sheetname: 表单名
        '''
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        '''读取Excel数据'''
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        # 获取第一行表头
        case = []
        title = [i.value for i in res[0]]

        # 遍历除第一行以外的其他行
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            case.append(dic)
            # print(case)
        # 返回读取后的数据
        return case

    def write_data(self, row, column, value):
        '''
        :param row: 写入的行
        :param column: 写入的列
        :param value: 写入的值
        :return:
        '''
        # 加载工作簿对象
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        # 写入数据
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)
